# inventory/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import Product, ProductImage, Vendor, VendorCost
from .forms import ProductForm
from django.db.models import Q
import random
from django.core.files.storage import default_storage
from django.contrib import messages
from import_export.resources import ModelResource
from django.utils.datastructures import MultiValueDictKeyError
from concurrent.futures import ThreadPoolExecutor
from django.conf import settings
import cv2
import os
import numpy as np
import pandas as pd
import tablib
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import zipfile

from rest_framework import status
from django.contrib.auth import authenticate
from rest_framework import generics, permissions
from rest_framework.response import Response
from django.contrib.auth.models import User
from rest_framework.views import APIView
from django.contrib.auth.models import User
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from .serializers import UserRegisterSerializer, UserLoginSerializer, UserSignupSerializer
  # Force garbage collection
class UserRegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserRegisterSerializer

class UserLoginView(generics.GenericAPIView):
    serializer_class = UserLoginSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        username = serializer.validated_data['username']
        password = serializer.validated_data['password']
        user = authenticate(username=username, password=password)

        if user is not None:
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            })
        return Response({'error': 'Invalid Credentials'}, status=400)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_detail_view(request):
    user = request.user
    return Response({
        'username': user.username,
        'email': user.email
    })
    
# Signup view
class SignupView(APIView):
    def post(self, request):
        serializer = UserSignupSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({"message": "User created successfully"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Helper function to create JWT tokens
def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }

# Login view
class LoginView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return Response({'error': 'Invalid username or password'}, status=status.HTTP_401_UNAUTHORIZED)

        if not user.check_password(password):
            return Response({'error': 'Invalid username or password'}, status=status.HTTP_401_UNAUTHORIZED)

        tokens = get_tokens_for_user(user)
        return Response(tokens, status=status.HTTP_200_OK)


def download_images(request, product_id):
    product = Product.objects.get(id=product_id)
    images = product.images.all()

    if not images:
        return HttpResponse("No images available for download.", content_type="text/plain")

    # Create a temporary zip file
    zip_filename = f"{product.title}_images.zip"
    zip_file_path = os.path.join(settings.MEDIA_ROOT, zip_filename)

    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
        for image in images:
            image_path = os.path.join(settings.MEDIA_ROOT, image.image.name)
            zipf.write(image_path, os.path.basename(image_path))

    # Serve the zip file as download
    response = HttpResponse(open(zip_file_path, 'rb'), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={zip_filename}'

    # Remove the temporary file after serving
    os.remove(zip_file_path)

    return response

def product_list(request):
    products = Product.objects.all().order_by('-created_at')  # Show latest products first

    # Check if it's a text search
    query = request.GET.get('q', '')
    if query:
        products = products.filter(
            Q(title__icontains=query) |
            Q(details__icontains=query) |
            Q(anime_name__icontains=query) |
            Q(character_name__icontains=query)
        ).prefetch_related('images')

    # Check if it's an image search
    if request.method == 'POST' and request.FILES.get('image'):
        image = request.FILES['image']
        query_image_path = save_uploaded_image(image)

        # Perform image search
        matched_product = search_by_image(query_image_path, products)
        if matched_product:
            products = [matched_product]  # Return only the matched product
        else:
            products = []  # No matching products found

    # Price filter
    max_price = request.GET.get('max_price')
    if max_price:
        products = products.filter(selling_price__lt=max_price)
        
    # Filter by Vendor
    vendor_id = request.GET.get('vendor')
    if vendor_id:
        products = products.filter(vendor_costs__vendor__id=vendor_id)

    # Sorting options
    sort_option = request.GET.get('sort')
    if sort_option == 'price_asc':
        products = products.order_by('selling_price')
    elif sort_option == 'price_desc':
        products = products.order_by('-selling_price')
    elif sort_option == 'size_asc':
        products = products.order_by('size')
    elif sort_option == 'size_desc':
        products = products.order_by('-size')

    return render(request, 'inventory/product_list.html', {'products': products, 'query': query, 'max_price': max_price, 'vendors': Vendor.objects.all()})

# Helper function to save the uploaded image temporarily
def save_uploaded_image(image):
    try:
        # Define the path for saving uploaded images in the 'uploads/' subfolder
        uploads_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
        
        # Create the 'uploads/' directory if it doesn't exist
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)

        # Generate the full path for the uploaded image
        query_image_path = os.path.join(uploads_dir, image.name)
        
        print(f"Saving image to: {query_image_path}")  # Debugging output
        
        # Save the uploaded image in chunks
        with open(query_image_path, 'wb+') as destination:
            for chunk in image.chunks():
                destination.write(chunk)

        print(f"Image successfully saved at: {query_image_path}")  # Debugging output
        return query_image_path

    except Exception as e:
        print(f"Error while saving image: {e}")
        return None

# # Image search function using OpenCV
# def search_by_image(query_image_path, products):
#     # Load the query image
#     query_img = cv2.imread(query_image_path, 0)
#     if query_img is None:
#         print(f"Query image {query_image_path} not found or failed to load.")
#         return None

#     sift = cv2.SIFT_create()  # Use SIFT to extract features
#     kp1, des1 = sift.detectAndCompute(query_img, None)
#     if des1 is None:
#         print("No features detected in query image.")
#         return None

#     print(f"Number of keypoints in query image: {len(kp1)}")

#     # Iterate over products and their images
#     for product in products:
#         if product.images.exists():
#             for product_image in product.images.all():
#                 # Construct the full path of the stored image
#                 stored_img_path = os.path.join(settings.MEDIA_ROOT, product_image.image.name)
                
#                 if not os.path.exists(stored_img_path):
#                     print(f"File {stored_img_path} does not exist.")
#                     continue

#                 stored_img = cv2.imread(stored_img_path, 0)  # Load the stored product image
#                 if stored_img is None:
#                     print(f"Error loading image: {stored_img_path}")
#                     continue

#                 kp2, des2 = sift.detectAndCompute(stored_img, None)
#                 if des2 is None:
#                     print(f"No features detected in stored image: {stored_img_path}")
#                     continue

#                 print(f"Matching with image: {stored_img_path}")

#                 # Match features
#                 bf = cv2.BFMatcher()
#                 matches = bf.knnMatch(des1, des2, k=2)

#                 # Filter good matches
#                 good_matches = [m for m, n in matches if m.distance < 0.75 * n.distance]

#                 print(f"Good matches found: {len(good_matches)}")

#                 # Adjust threshold as needed
#                 if len(good_matches) > 90:  # Adjust threshold based on your data
#                     print(f"Product matched: {product.title}")
#                     return product  # Return the matched product

#     print("No matching product found.")
#     return None


# Image search function using OpenCV and optimizations
def search_by_image(query_image_path, products):
    # Load the query image
    query_img = cv2.imread(query_image_path, 0)
    if query_img is None:
        print(f"Query image {query_image_path} not found or failed to load.")
        return None

    # SIFT with a limited number of keypoints
    sift = cv2.SIFT_create(nfeatures=500)  # Limit keypoints to 500 for consistency
    kp1, des1 = sift.detectAndCompute(query_img, None)
    if des1 is None:
        print("No features detected in query image.")
        return None

    print(f"Number of keypoints in query image: {len(kp1)}")

    # Initialize FLANN-based matcher (Fast Approximate Nearest Neighbors)
    FLANN_INDEX_KDTREE = 1
    index_params = dict(algorithm=FLANN_INDEX_KDTREE, trees=5)
    search_params = dict(checks=50)  # Increase checks for better precision
    flann = cv2.FlannBasedMatcher(index_params, search_params)

    def match_product_image(product, product_image):
        stored_img_path = os.path.join(settings.MEDIA_ROOT, product_image.image.name)

        if not os.path.exists(stored_img_path):
            print(f"File {stored_img_path} does not exist.")
            return None, 0

        stored_img = cv2.imread(stored_img_path, 0)  # Load the stored product image
        if stored_img is None:
            print(f"Error loading image: {stored_img_path}")
            return None, 0

        kp2, des2 = sift.detectAndCompute(stored_img, None)
        if des2 is None:
            print(f"No features detected in stored image: {stored_img_path}")
            return None, 0

        # Match features using FLANN-based matcher
        matches = flann.knnMatch(des1, des2, k=2)

        # Filter good matches
        good_matches = [m for m, n in matches if m.distance < 0.7 * n.distance]

        return product, len(good_matches)

    # Parallelize image matching with ThreadPoolExecutor
    with ThreadPoolExecutor() as executor:
        futures = []
        for product in products:
            if product.images.exists():
                for product_image in product.images.all():
                    futures.append(executor.submit(match_product_image, product, product_image))

        results = [future.result() for future in futures]

    # Find the best matching product based on good matches
    best_product, max_matches = max(results, key=lambda x: x[1])

    print(f"Good matches found: {max_matches}")

    # Use dynamic threshold for a match
    if max_matches > 100:  # Adjust based on data
        print(f"Product matched: {best_product.title}")
        return best_product

    print("No matching product found.")
    return None

def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    images = product.images.all()  # Assuming images is a related name for product images
    vendors = VendorCost.objects.filter(product=product)  # Get all vendor costs for this product

    return render(request, 'inventory/product_detail.html', {
        'product': product,
        'images': images,
        'vendors': vendors,
    })

def add_product(request):
    if request.method == 'POST':
        # Get the form data
        title = request.POST['title']
        details = request.POST['details']
        anime_name = request.POST['anime_name']
        character_name = request.POST['character_name']
        dimensions = request.POST['dimensions']
        selling_price = request.POST.get('selling_price')
        size = request.POST['size']  # Size in the format "10cm x 50cm"
        weight = request.POST['weight']
        additional_charges = request.POST.get('additional_charges', 0)
        in_stock = bool(request.POST.get('in_stock', False))
        pre_order = bool(request.POST.get('pre_order', False))

        # Get vendors and their cost prices from the dropdown
        vendor_ids = request.POST.getlist('vendor_ids[]')  # Get selected vendor IDs
        cost_prices = request.POST.getlist('cost_price[]')  # Get corresponding cost prices

        # Handle the video upload
        video = request.FILES.get('video')

        # Handle multiple image uploads
        images = request.FILES.getlist('pictures')

        # Check for duplicate products (search by image)
        duplicate_product = None
        if images:
            # Use the first image for duplicate detection
            first_image = images[0]
            duplicate_product = search_by_image(first_image.temporary_file_path(), Product.objects.all())

        if duplicate_product:
            # Duplicate found, ask if the user wants to update or save as new
            messages.warning(request, "A product with the same image already exists. Do you want to update the existing product or save this as a new one?")
            return render(request, 'inventory/confirm_save.html', {'duplicate_product': duplicate_product, 'new_product_data': request.POST, 'new_images': images})

        else:
            # Create and save the product
            product = Product.objects.create(
                title=title,
                details=details,
                anime_name=anime_name,
                character_name=character_name,
                selling_price=selling_price,
                dimensions=dimensions,
                size=size,  # Save the new size field
                weight=weight,
                additional_charges=additional_charges if additional_charges else None,
                in_stock=in_stock,
                pre_order=pre_order,
                video=video  # Save the uploaded video
            )

            # Create VendorCost instances
            for vendor_id, cost_price in zip(vendor_ids, cost_prices):
                if cost_price is not None:  # Only create VendorCost if cost_price is valid
                    vendor = Vendor.objects.get(id=vendor_id)
                    VendorCost.objects.create(
                        product=product,
                        vendor=vendor,
                        cost_price=cost_price
                    )

            # Handle multiple image uploads and rename them
            for image in images:
                # Generate a unique 4-digit number
                unique_id = str(random.randint(1000, 9999))

                # Define the new file name
                extension = os.path.splitext(image.name)[1]  # Get the file extension (e.g., .jpg, .png)
                new_filename = f"{title}_{unique_id}{extension}"

                # Save the renamed image
                image.name = new_filename  # Rename the file
                ProductImage.objects.create(product=product, image=image)

            messages.success(request, "Product added successfully!")
            return redirect('product_list')

    # Fetch all vendors to display in the dropdown
    vendors = Vendor.objects.all()
    return render(request, 'inventory/add_product.html', {'vendors': vendors})


def update_product(request, pk):
    product = get_object_or_404(Product, id=pk)

    if request.method == 'POST':
        # Get the form data from the request
        title = request.POST['title']
        details = request.POST['details']
        anime_name = request.POST['anime_name']
        character_name = request.POST['character_name']
        dimensions = request.POST['dimensions']
        selling_price = request.POST.get('selling_price')
        size = request.POST['size']
        weight = request.POST['weight']
        additional_charges = request.POST.get('additional_charges', 0)
        in_stock = bool(request.POST.get('in_stock', False))
        pre_order = bool(request.POST.get('pre_order', False))

        # Get vendors and their cost prices from the form
        vendor_ids = request.POST.getlist('vendor_ids[]')  # Get selected vendor IDs
        cost_prices = request.POST.getlist('cost_price[]')
        
        

        # Handle video file update
        video = request.FILES.get('video')

        # Update the product fields
        product.title = title
        product.details = details
        product.anime_name = anime_name
        product.character_name = character_name
        product.selling_price = selling_price
        product.dimensions = dimensions
        product.size = size
        product.weight = weight
        product.additional_charges = additional_charges if additional_charges else None
        product.in_stock = in_stock
        product.pre_order = pre_order
        if video:
            product.video = video  # Update video if provided
        product.save()

        # Update VendorCost instances
        # # Remove old vendor cost data
        product.vendor_costs.all().delete() 
        # Create VendorCost instances
        for vendor_id, cost_price in zip(vendor_ids, cost_prices):
            if cost_price is not None:  # Only create VendorCost if cost_price is valid
                vendor = Vendor.objects.get(id=vendor_id)
                VendorCost.objects.create(
                    product=product,
                    vendor=vendor,
                    cost_price=cost_price
                )

        # Handle image updates
        images = request.FILES.getlist('pictures')
        if images:
            product.images.all().delete()  # Remove old images
            for image in images:
                unique_id = str(random.randint(1000, 9999))
                extension = os.path.splitext(image.name)[1]
                new_filename = f"{title}_{unique_id}{extension}"
                image.name = new_filename
                ProductImage.objects.create(product=product, image=image)

        messages.success(request, "Product updated successfully!")
        return redirect('product_list')

    # Fetch all vendors and prefill form with product data
    vendors = Vendor.objects.all()
    vendor_costs = {vc.vendor.id: vc.cost_price for vc in VendorCost.objects.filter(product=pk)}
    print(vendors, vendor_costs)

    return render(request, 'inventory/update_product.html', {
        'product': product,
        'vendors': vendors,
        'vendor_costs': vendor_costs
    })

def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        # Delete associated product images
        ProductImage.objects.filter(product=product).delete()

        # Delete associated vendor costs
        VendorCost.objects.filter(product=product).delete()

        # Finally, delete the product itself
        product.delete()

        # Add a success message and redirect to product list
        messages.success(request, f'Product "{product.title}" and all associated data deleted successfully!')
        return redirect('product_list')

    return render(request, 'inventory/delete_product.html', {'product': product})



def export_to_excel(request):
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers for the Excel sheet
    headers = ['Title', 'Details', 'Anime Name', 'Character Name', 'Vendor Name', 'Vendor Phone', 'Cost Price', 'Image']
    ws.append(headers)

    # Set column widths to make space for the images
    ws.column_dimensions['A'].width = 20  # Title
    ws.column_dimensions['B'].width = 30  # Details
    ws.column_dimensions['C'].width = 20  # Anime Name
    ws.column_dimensions['D'].width = 20  # Character Name
    ws.column_dimensions['E'].width = 20  # Vendor Name
    ws.column_dimensions['F'].width = 15  # Vendor Phone
    ws.column_dimensions['G'].width = 15  # Cost Price
    ws.column_dimensions['H'].width = 20  # Image

    # Fetch data from your models
    for product in Product.objects.all():
        vendor_costs = product.vendor_costs.all()  # Fetch associated vendor costs

        # Only one image per product, fetch the first one
        product_image = product.images.first()  # Assumes only one image is needed

        for vendor_cost in vendor_costs:
            # Create the row of data
            row = [
                product.title,
                product.details,
                product.anime_name,
                product.character_name,
                vendor_cost.vendor.name,  # Vendor Name
                vendor_cost.vendor.phone_number,  # Vendor Phone
                vendor_cost.cost_price,  # Cost Price
            ]

            ws.append(row)

            # If the product has an image, insert it into the corresponding cell
            if product_image:
                try:
                    image_path = default_storage.path(product_image.image.name)

                    # Debugging: Print the image path to verify it's correct
                    print(f"Image Path: {image_path}")

                    # Check if file exists at the given path
                    if os.path.exists(image_path):
                        img = ExcelImage(image_path)

                        # Optionally resize the image (width x height in pixels)
                        img.width = 100
                        img.height = 100

                        # Add image to the corresponding cell in column H
                        img_cell = f'H{ws.max_row}'  # H-column for the image
                        ws.add_image(img, img_cell)
                    else:
                        # Debugging: Image file not found
                        print(f"Image file not found: {image_path}")

                except Exception as e:
                    # Debugging: Print any errors during image handling
                    print(f"Error loading image: {e}")

    # Create an HTTP response with the appropriate content type for Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


def import_from_excel(request):
    if request.method == 'POST':
        file = request.FILES.get('file')

        if file and (file.name.endswith('.xlsx') or file.name.endswith('.xls')):
            try:
                # Load the Excel file
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Loop through each row in the sheet (assuming row 1 is header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Ensure row has at least 11 values to unpack
                    if len(row) < 11:
                        continue  # Skip rows that do not have enough data
                    
                    title, details, anime_name, character_name, selling_price, dimensions, weight, size, additional_charges, vendor_name, cost_price, *image_filenames = row[:12]

                    # Create or update vendor
                    vendor, _ = Vendor.objects.get_or_create(name=vendor_name)

                    # Create new product or update if it already exists
                    product, _ = Product.objects.update_or_create(
                        title=title,
                        defaults={
                            'details': details,
                            'anime_name': anime_name,
                            'character_name': character_name,
                            'selling_price': selling_price,
                            'dimensions': dimensions,
                            'weight': weight,
                            'size': size,
                            'additional_charges': additional_charges,
                        }
                    )

                    # Create or update vendor cost associated with the product
                    VendorCost.objects.update_or_create(
                        product=product,
                        vendor=vendor,
                        defaults={'cost_price': cost_price}
                    )

                    # Process images
                    for image_filename in image_filenames:
                        if image_filename:  # Check if the filename is not empty
                            # Construct full image path
                            image_path = os.path.join(settings.MEDIA_ROOT, 'product_images', image_filename)

                            # Check if the image already exists in the media folder
                            if os.path.isfile(image_path):
                                # Only create ProductImage if it doesn't already exist for this product
                                ProductImage.objects.get_or_create(product=product, image=image_filename)

                messages.success(request, "Products and vendors imported successfully!")
                return redirect('product_list')

            except Exception as e:
                return render(request, 'inventory/import_products.html', {
                    'error_message': f"Error processing file: {str(e)}"
                })
        else:
            return render(request, 'inventory/import_products.html', {
                'error_message': "Invalid file format. Please upload a .xlsx or .xls file."
            })

    return render(request, 'inventory/import_products.html')




def bulk_action(request):
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        action = request.POST.get('action')

        if not product_ids:
            messages.error(request, "No products selected.")
            return redirect('product_list')

        if action == 'delete':
            Product.objects.filter(id__in=product_ids).delete()
            messages.success(request, "Selected products deleted successfully.")

        elif action == 'export':
            # Create a new dataset for export
            dataset = tablib.Dataset()
            dataset.headers = [
                'Title', 'Details', 'Anime Name', 'Character Name',
                'Selling Price', 'Dimensions', 'Weight',
                'Size', 'Additional Charges', 'Vendor', 'Cost Price', 'Images'
            ]

            # Add the selected products to the dataset
            for product in Product.objects.filter(id__in=product_ids):
                # Collect vendor information
                vendor_costs = product.vendor_costs.all()
                for vendor_cost in vendor_costs:
                    dataset.append([
                        product.title,
                        product.details,
                        product.anime_name,
                        product.character_name,
                        product.selling_price,
                        product.dimensions,
                        product.weight,
                        product.size,
                        product.additional_charges,
                        vendor_cost.vendor.name,  # Vendor Name
                        vendor_cost.cost_price,  # Cost Price
                        # Join all image URLs
                        ', '.join([img.image.url for img in product.images.all()])
                    ])

            # Create an HTTP response with the correct content type for Excel
            response = HttpResponse(
                dataset.export('xlsx'), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="selected_products.xlsx"'
            return response

        else:
            messages.error(request, "Invalid action.")


# Create a new vendor
def add_vendor(request):
    if request.method == 'POST':
        name = request.POST['name']
        phone_number = request.POST.get('phone_number', None)
        address = request.POST.get('address', None)
        payment_upi_id = request.POST.get('payment_upi_id', None)

        vendor = Vendor.objects.create(
            name=name,
            phone_number=phone_number,
            address=address,
            payment_upi_id=payment_upi_id
        )

        messages.success(request, 'Vendor added successfully!')
        return redirect('vendor_list')  # Redirect to vendor list page

    return render(request, 'inventory/add_vendor.html')


# Update an existing vendor
def update_vendor(request, vendor_id):
    vendor = get_object_or_404(Vendor, id=vendor_id)

    if request.method == 'POST':
        vendor.name = request.POST['name']
        vendor.phone_number = request.POST.get('phone_number', None)
        vendor.address = request.POST.get('address', None)
        vendor.payment_upi_id = request.POST.get('payment_upi_id', None)
        vendor.save()

        messages.success(request, 'Vendor updated successfully!')
        return redirect('vendor_list')  # Redirect to vendor list page

    return render(request, 'inventory/update_vendor.html', {'vendor': vendor})


# Delete a vendor
def delete_vendor(request, vendor_id):
    vendor = get_object_or_404(Vendor, id=vendor_id)
    
    if request.method == 'POST':
        vendor.delete()
        messages.success(request, 'Vendor deleted successfully!')
        return redirect('vendor_list')

    return render(request, 'inventory/delete_vendor.html', {'vendor': vendor})


# List all vendors
def list_vendors(request):
    vendors = Vendor.objects.all()
    return render(request, 'inventory/vendor_list.html', {'vendors': vendors})